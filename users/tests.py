from openpyxl import Workbook

wb = Workbook()
ws = wb.active

for i in range(1,10):
    ws.cell(row=i,column=1,value="first")
    ws.cell(row=i,column=2,value="second")
    ws.cell(row=i,column=3,value="third")

wb.save("My_test.xlsx")
